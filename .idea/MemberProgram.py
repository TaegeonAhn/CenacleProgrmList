import csv
import time
import os
import openpyxl

start = time.time()

file_path_list =[]

def Search(dirname):
    try:
        filenames = os.listdir(dirname)
        for filename in filenames:
            full_filename = os.path.join(dirname, filename)
            if os.path.isdir(full_filename):
                search(full_filename)
            else:
                ext = os.path.splitext(full_filename)[-1]
                if ext == '.csv':
                    full_filename = full_filename.replace("/", "\\")
                    file_path_list.append(full_filename)
    except PermissionError:
        pass

Search('D:/10.실행파일_win') #txt 파일 읽기

dic_result = {}

for file_path in file_path_list:
    file_path_tokens = file_path.split("_")
    memberName = file_path_tokens[-2]

    freader = open(file_path,'r',encoding='utf-8')
    reader = csv.reader(freader)

    i = 0
    vender_index = 0
    for line in reader:
        file_name = line[1]

        if i == 0:
            for cellindex in range(len(line)):
                if 'Vendor' in line[cellindex]:
                    vender_index = cellindex
        else:
            if file_name == '' or len(file_name) < 1:
                continue

            if file_name in dic_result:
                dic_result[file_name] += "/" + memberName

            else:
                dic_result[file_name] = line[vender_index] + "+" + memberName

        i += 1

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "sheet1"
sheet.cell(row =1 , column=1).value = '파일명'
sheet.cell(row =1 , column=2).value = '공급업체명'
sheet.cell(row =1 , column=3).value = '사용자명'

i = 2

for key in dic_result.keys():
    dic_tokens = dic_result[key].split("+")

    sheet.cell(row = i, column=1).value = key
    sheet.cell(row = i, column=2).value = dic_tokens[0]
    sheet.cell(row = i, column=3).value = dic_tokens[1]
    i += 1

result_file_path = 'D:/10.실행파일_win/product_win_세나클_20200714.xlsx'
wb.save(result_file_path)



freader.close()





